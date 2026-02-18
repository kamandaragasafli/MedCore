"""
Master Admin Panel Views
- Dashboard with platform analytics
- Company management
- User management
- Impersonation/Tenant switching
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db.models import Count, Q, Sum
from django.utils import timezone
from datetime import timedelta
from django.contrib.auth.models import User
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from subscription.models import Company, Subscription, UserProfile, SubscriptionPlan, Notification, NotificationTemplate
from .decorators import superuser_required
from django.views.decorators.http import require_http_methods


@superuser_required
def master_dashboard(request):
    """
    Master Admin Dashboard
    - Platform-wide analytics
    - Quick stats
    - Recent activity
    """
    
    # Company statistics
    total_companies = Company.objects.count()
    active_companies = Company.objects.filter(
        subscriptions__status='active',
        subscriptions__end_date__gte=timezone.now()
    ).distinct().count()
    
    inactive_companies = total_companies - active_companies
    
    # User statistics
    total_users = User.objects.count()
    total_company_users = UserProfile.objects.count()
    superusers = User.objects.filter(is_superuser=True).count()
    
    # Recent companies (last 30 days)
    thirty_days_ago = timezone.now() - timedelta(days=30)
    recent_companies = Company.objects.filter(
        created_at__gte=thirty_days_ago
    ).count()
    
    # Subscription statistics
    subscription_plans = SubscriptionPlan.objects.annotate(
        subscriber_count=Count('subscription', filter=Q(subscription__status='active'))
    )
    
    # Recent companies list
    latest_companies = Company.objects.order_by('-created_at')[:10]
    
    # Active subscriptions
    active_subscriptions = Subscription.objects.filter(
        status='active',
        end_date__gte=timezone.now()
    ).select_related('company', 'plan').order_by('-start_date')[:10]
    
    context = {
        'total_companies': total_companies,
        'active_companies': active_companies,
        'trial_companies': 0,  # No longer used
        'inactive_companies': inactive_companies,
        'total_users': total_users,
        'total_company_users': total_company_users,
        'superusers': superusers,
        'recent_companies': recent_companies,
        'subscription_plans': subscription_plans,
        'latest_companies': latest_companies,
        'active_subscriptions': active_subscriptions,
    }
    
    return render(request, 'master_admin/dashboard.html', context)


@superuser_required
def company_list(request):
    """
    List all companies with detailed information
    - Search and filter
    - Status indicators
    - Quick actions
    """
    
    # Get all companies with related data
    companies = Company.objects.prefetch_related(
        'subscriptions',
        'user_profiles__user'
    ).order_by('-created_at')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        companies = companies.filter(
            Q(name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(phone__icontains=search_query)
        )
    
    # Status filter
    status_filter = request.GET.get('status', '')
    if status_filter == 'active':
        companies = companies.filter(
            subscriptions__status='active',
            subscriptions__end_date__gte=timezone.now()
        ).distinct()
    elif status_filter == 'inactive':
        # Companies without active subscriptions
        active_companies = Company.objects.filter(
            subscriptions__status='active',
            subscriptions__end_date__gte=timezone.now()
        ).values_list('id', flat=True)
        companies = companies.exclude(id__in=active_companies)
    
    # Annotate with stats
    companies_data = []
    for company in companies:
        # Get active subscription
        active_sub = company.active_subscription
        
        # Get user count
        user_count = company.user_profiles.filter(is_active=True).count()
        
        companies_data.append({
            'company': company,
            'subscription': active_sub,
            'user_count': user_count,
            'status': 'active' if active_sub else 'inactive'
        })
    
    context = {
        'companies_data': companies_data,
        'search_query': search_query,
        'status_filter': status_filter,
    }
    
    return render(request, 'master_admin/company_list.html', context)


@superuser_required
def company_detail(request, company_id):
    """
    Detailed view of a specific company
    - Company info
    - Subscription history
    - Users
    - Statistics
    - Doctors, Finances, Medications (from tenant database)
    """
    
    company = get_object_or_404(Company, id=company_id)
    
    # Get all subscriptions
    subscriptions = company.subscriptions.all().order_by('-start_date')
    
    # Get all users
    users = company.user_profiles.select_related('user').all()
    
    # Get active subscription
    active_subscription = company.active_subscription
    
    # Get tenant database data
    doctors = []
    drugs = []
    total_debt = 0
    doctors_count = 0
    drugs_count = 0
    active_doctors_count = 0
    tenant_db_error = None
    
    if company.db_name:
        from subscription.db_router import set_tenant_db, clear_tenant_db
        from doctors.models import Doctor
        from drugs.models import Drug
        from django.db.models import Sum
        from django.db import OperationalError, DatabaseError
        
        try:
            # Set tenant database context
            set_tenant_db(company.db_name)
            
            # Check if tables exist by trying to count
            try:
                # Get doctors with related data
                doctors = list(Doctor.objects.select_related(
                    'region', 'city', 'clinic', 'ixtisas'
                ).all().order_by('-created_at')[:50])  # Limit to 50 for display
                
                doctors_count = Doctor.objects.count()
                active_doctors_count = Doctor.objects.filter(is_active=True).count()
                
                # Calculate total debt
                debt_sum = Doctor.objects.aggregate(
                    total=Sum('yekun_borc')
                )
                total_debt = float(debt_sum['total'] or 0)
            except (OperationalError, DatabaseError) as e:
                # Tables don't exist - migrations not run
                tenant_db_error = "Verilənlər bazası cədvəlləri yoxdur. Migrasiyaları icra edin."
                doctors = []
                doctors_count = 0
                active_doctors_count = 0
                total_debt = 0
            
            # Get drugs/medications
            try:
                drugs = list(Drug.objects.filter(is_active=True).order_by('ad')[:50])  # Limit to 50 for display
                drugs_count = Drug.objects.filter(is_active=True).count()
            except (OperationalError, DatabaseError):
                # Drugs table doesn't exist
                drugs = []
                drugs_count = 0
            
        except Exception as e:
            # If tenant database doesn't exist or has other issues
            tenant_db_error = f"Verilənlər bazasına giriş xətası: {str(e)}"
        finally:
            clear_tenant_db()
    
    context = {
        'company': company,
        'subscriptions': subscriptions,
        'users': users,
        'active_subscription': active_subscription,
        'user_count': users.count(),
        'doctors': doctors,
        'doctors_count': doctors_count,
        'active_doctors_count': active_doctors_count,
        'drugs': drugs,
        'drugs_count': drugs_count,
        'total_debt': total_debt,
        'tenant_db_error': tenant_db_error,
    }
    
    return render(request, 'master_admin/company_detail.html', context)


@superuser_required
def switch_to_company(request, company_id):
    """
    Impersonate / Switch to company's system
    - Store current state
    - Switch database context
    - Redirect to company dashboard
    """
    
    company = get_object_or_404(Company, id=company_id)
    
    # Store master admin state in session
    request.session['master_admin_mode'] = True
    request.session['master_admin_original_company'] = request.company.id if hasattr(request, 'company') and request.company else None
    request.session['impersonating_company_id'] = company.id
    request.session['impersonating_company_name'] = company.name
    
    # Force set company context
    request.company = company
    
    messages.success(
        request,
        f'Siz indi {company.name} şirkətinin sisteminə daxil oldunuz. '
        f'Master Admin panelinə qayıtmaq üçün "Master Panelə Qayıt" düyməsinə basın.'
    )
    
    return redirect('core:dashboard')


@superuser_required
def exit_impersonation(request):
    """
    Exit company impersonation and return to Master Admin Panel
    """
    
    company_name = request.session.get('impersonating_company_name', 'şirkət')
    
    # Clear impersonation session data
    request.session.pop('master_admin_mode', None)
    request.session.pop('master_admin_original_company', None)
    request.session.pop('impersonating_company_id', None)
    request.session.pop('impersonating_company_name', None)
    
    messages.success(
        request,
        f'{company_name} sistemindən çıxış edildi. Master Admin panelinə qayıtdınız.'
    )
    
    return redirect('master_admin:dashboard')


@superuser_required
def platform_analytics(request):
    """
    Platform-wide analytics and statistics
    - User growth
    - Company growth
    - Revenue analytics
    - Usage statistics
    """
    
    # Time periods
    today = timezone.now()
    last_7_days = today - timedelta(days=7)
    last_30_days = today - timedelta(days=30)
    last_90_days = today - timedelta(days=90)
    
    # Company growth
    companies_last_7 = Company.objects.filter(created_at__gte=last_7_days).count()
    companies_last_30 = Company.objects.filter(created_at__gte=last_30_days).count()
    companies_last_90 = Company.objects.filter(created_at__gte=last_90_days).count()
    
    # User growth
    users_last_7 = UserProfile.objects.filter(joined_at__gte=last_7_days).count()
    users_last_30 = UserProfile.objects.filter(joined_at__gte=last_30_days).count()
    users_last_90 = UserProfile.objects.filter(joined_at__gte=last_90_days).count()
    
    # Subscription distribution
    subscription_distribution = {}
    for plan in SubscriptionPlan.objects.all():
        count = Subscription.objects.filter(
            plan=plan,
            status='active',
            end_date__gte=today
        ).count()
        subscription_distribution[plan.name] = count
    
    # Company status distribution
    total_companies = Company.objects.count()
    active_count = Company.objects.filter(
        subscriptions__status='active',
        subscriptions__end_date__gte=today
    ).distinct().count()
    inactive_count = total_companies - active_count
    
    # Monthly company registrations (last 12 months)
    monthly_registrations = []
    for i in range(11, -1, -1):
        month_start = today - timedelta(days=30*i)
        month_end = today - timedelta(days=30*(i-1)) if i > 0 else today
        count = Company.objects.filter(
            created_at__gte=month_start,
            created_at__lt=month_end
        ).count()
        monthly_registrations.append({
            'month': month_start.strftime('%b %Y'),
            'count': count
        })
    
    context = {
        'companies_last_7': companies_last_7,
        'companies_last_30': companies_last_30,
        'companies_last_90': companies_last_90,
        'users_last_7': users_last_7,
        'users_last_30': users_last_30,
        'users_last_90': users_last_90,
        'subscription_distribution': subscription_distribution,
        'active_count': active_count,
        'trial_count': 0,  # No longer used
        'inactive_count': inactive_count,
        'monthly_registrations': monthly_registrations,
    }
    
    return render(request, 'master_admin/analytics.html', context)


@superuser_required
def user_management(request):
    """
    Manage all platform users
    - Search users
    - View user details
    - Deactivate/activate users
    """
    
    users = UserProfile.objects.select_related('user', 'company').order_by('-joined_at')
    
    # Search
    search_query = request.GET.get('search', '')
    if search_query:
        users = users.filter(
            Q(user__username__icontains=search_query) |
            Q(user__email__icontains=search_query) |
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(company__name__icontains=search_query)
        )
    
    # Role filter
    role_filter = request.GET.get('role', '')
    if role_filter:
        users = users.filter(role=role_filter)
    
    context = {
        'users': users,
        'search_query': search_query,
        'role_filter': role_filter,
    }
    
    return render(request, 'master_admin/user_management.html', context)


@superuser_required
@require_http_methods(["GET", "POST"])
def send_notification(request, company_id=None):
    """
    Send notification to company/companies
    - Single company or multiple companies
    - Different notification types
    """
    
    if request.method == 'POST':
        # Get form data
        company_ids = request.POST.getlist('companies')
        title = request.POST.get('title', '').strip()
        message = request.POST.get('message', '').strip()
        notification_type = request.POST.get('notification_type', 'info')
        is_important = request.POST.get('is_important') == 'on'
        action_url = request.POST.get('action_url', '').strip()
        action_text = request.POST.get('action_text', '').strip()
        
        # Validation
        if not company_ids:
            messages.error(request, 'Ən azı bir şirkət seçilməlidir.')
            return redirect('master_admin:send_notification')
        
        if not title or not message:
            messages.error(request, 'Başlıq və mesaj mütləqdir.')
            return redirect('master_admin:send_notification')
        
        # Send notification to selected companies
        sent_count = 0
        for company_id in company_ids:
            try:
                company = Company.objects.get(id=company_id)
                Notification.objects.create(
                    company=company,
                    title=title,
                    message=message,
                    notification_type=notification_type,
                    is_important=is_important,
                    action_url=action_url if action_url else None,
                    action_text=action_text if action_text else None,
                    created_by=request.user
                )
                sent_count += 1
            except Company.DoesNotExist:
                continue
        
        messages.success(
            request, 
            f'{sent_count} şirkətə bildiriş göndərildi.'
        )
        
        # Redirect based on context
        if len(company_ids) == 1:
            return redirect('master_admin:company_detail', company_id=company_ids[0])
        else:
            return redirect('master_admin:send_notification')
    
    # GET request - show form
    companies = Company.objects.filter(is_active=True).order_by('name')
    
    # If company_id is provided, pre-select that company
    selected_company = None
    if company_id:
        try:
            selected_company = Company.objects.get(id=company_id)
        except Company.DoesNotExist:
            pass
    
    # Get notification templates
    templates = NotificationTemplate.objects.filter(is_active=True).order_by('name')
    
    # Check if template_id is provided in query params
    selected_template = None
    template_id = request.GET.get('template_id')
    if template_id:
        try:
            selected_template = NotificationTemplate.objects.get(id=template_id, is_active=True)
        except NotificationTemplate.DoesNotExist:
            pass
    
    context = {
        'companies': companies,
        'selected_company': selected_company,
        'templates': templates,
        'selected_template': selected_template,
    }
    
    return render(request, 'master_admin/send_notification.html', context)


@superuser_required
def export_company_doctors_excel(request, company_id):
    """
    Export all doctors from a company to Excel file
    """
    company = get_object_or_404(Company, id=company_id)
    
    if not company.db_name:
        messages.error(request, 'Bu şirkətin verilənlər bazası yoxdur.')
        return redirect('master_admin:company_detail', company_id=company_id)
    
    from subscription.db_router import set_tenant_db, clear_tenant_db
    from doctors.models import Doctor
    
    try:
        # Set tenant database context
        set_tenant_db(company.db_name)
        
        # Get all doctors
        doctors = Doctor.objects.select_related(
            'region', 'city', 'clinic', 'ixtisas'
        ).all().order_by('code')
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Həkimlər"
        
        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        cell_font = Font(name='Arial', size=11)
        cell_alignment = Alignment(horizontal='left', vertical='center')
        number_alignment = Alignment(horizontal='right', vertical='center')
        
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Define column headers
        headers = [
            ('Kod', 12),
            ('Ad Soyad', 25),
            ('İxtisas', 20),
            ('Bölgə', 15),
            ('Şəhər', 15),
            ('Klinika', 30),
            ('Telefon', 18),
            ('Email', 25),
            ('Cinsiyyət', 12),
            ('Kateqoriya', 12),
            ('Dərəcə', 10),
            ('Əvvəlki Borc', 15),
            ('Hesablanmış', 15),
            ('Silinən', 15),
            ('Yekun Borc', 15),
            ('Qeydiyyat Tarixi', 18),
            ('Status', 12),
        ]
        
        # Write headers
        for col_num, (header, width) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
            ws.column_dimensions[cell.column_letter].width = width
        
        # Write data
        for row_num, doctor in enumerate(doctors, 2):
            ws.cell(row=row_num, column=1, value=doctor.code).font = cell_font
            ws.cell(row=row_num, column=1).alignment = cell_alignment
            ws.cell(row=row_num, column=1).border = border
            
            ws.cell(row=row_num, column=2, value=doctor.ad).font = cell_font
            ws.cell(row=row_num, column=2).alignment = cell_alignment
            ws.cell(row=row_num, column=2).border = border
            
            ws.cell(row=row_num, column=3, value=doctor.ixtisas.name if doctor.ixtisas else '-').font = cell_font
            ws.cell(row=row_num, column=3).alignment = cell_alignment
            ws.cell(row=row_num, column=3).border = border
            
            ws.cell(row=row_num, column=4, value=doctor.region.name if doctor.region else '-').font = cell_font
            ws.cell(row=row_num, column=4).alignment = cell_alignment
            ws.cell(row=row_num, column=4).border = border
            
            ws.cell(row=row_num, column=5, value=doctor.city.name if doctor.city else '-').font = cell_font
            ws.cell(row=row_num, column=5).alignment = cell_alignment
            ws.cell(row=row_num, column=5).border = border
            
            ws.cell(row=row_num, column=6, value=doctor.clinic.name if doctor.clinic else '-').font = cell_font
            ws.cell(row=row_num, column=6).alignment = cell_alignment
            ws.cell(row=row_num, column=6).border = border
            
            ws.cell(row=row_num, column=7, value=doctor.telefon or '-').font = cell_font
            ws.cell(row=row_num, column=7).alignment = cell_alignment
            ws.cell(row=row_num, column=7).border = border
            
            ws.cell(row=row_num, column=8, value=doctor.email or '-').font = cell_font
            ws.cell(row=row_num, column=8).alignment = cell_alignment
            ws.cell(row=row_num, column=8).border = border
            
            ws.cell(row=row_num, column=9, value=doctor.get_gender_display()).font = cell_font
            ws.cell(row=row_num, column=9).alignment = cell_alignment
            ws.cell(row=row_num, column=9).border = border
            
            ws.cell(row=row_num, column=10, value=doctor.get_category_display()).font = cell_font
            ws.cell(row=row_num, column=10).alignment = cell_alignment
            ws.cell(row=row_num, column=10).border = border
            
            ws.cell(row=row_num, column=11, value=doctor.get_degree_display()).font = cell_font
            ws.cell(row=row_num, column=11).alignment = cell_alignment
            ws.cell(row=row_num, column=11).border = border
            
            ws.cell(row=row_num, column=12, value=float(doctor.evvelki_borc or 0)).font = cell_font
            ws.cell(row=row_num, column=12).alignment = number_alignment
            ws.cell(row=row_num, column=12).border = border
            ws.cell(row=row_num, column=12).number_format = '#,##0.00'
            
            ws.cell(row=row_num, column=13, value=float(doctor.hesablanmish_miqdar or 0)).font = cell_font
            ws.cell(row=row_num, column=13).alignment = number_alignment
            ws.cell(row=row_num, column=13).border = border
            ws.cell(row=row_num, column=13).number_format = '#,##0.00'
            
            ws.cell(row=row_num, column=14, value=float(doctor.silinen_miqdar or 0)).font = cell_font
            ws.cell(row=row_num, column=14).alignment = number_alignment
            ws.cell(row=row_num, column=14).border = border
            ws.cell(row=row_num, column=14).number_format = '#,##0.00'
            
            # Yekun Borc with color coding
            debt_cell = ws.cell(row=row_num, column=15, value=float(doctor.yekun_borc or 0))
            debt_cell.font = Font(name='Arial', size=11, color='DC2626' if doctor.yekun_borc > 0 else ('22C55E' if doctor.yekun_borc < 0 else '000000'))
            debt_cell.alignment = number_alignment
            debt_cell.border = border
            debt_cell.number_format = '#,##0.00'
            
            ws.cell(row=row_num, column=16, value=doctor.created_at.strftime('%d.%m.%Y') if doctor.created_at else '-').font = cell_font
            ws.cell(row=row_num, column=16).alignment = cell_alignment
            ws.cell(row=row_num, column=16).border = border
            
            # Status with color coding
            status_cell = ws.cell(row=row_num, column=17, value='Aktiv' if doctor.is_active else 'Deaktiv')
            status_cell.font = cell_font
            status_cell.alignment = cell_alignment
            status_cell.border = border
            status_cell.fill = PatternFill(start_color='DCfCE7' if doctor.is_active else 'FEE2E2', end_color='DCfCE7' if doctor.is_active else 'FEE2E2', fill_type='solid')
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Generate filename
        from datetime import datetime
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{company.name}_Hekimler_{timestamp}.xlsx"
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        
        return response
        
    except Exception as e:
        messages.error(request, f'Excel faylı yaradılarkən xəta baş verdi: {str(e)}')
        return redirect('master_admin:company_detail', company_id=company_id)
    finally:
        clear_tenant_db()


@superuser_required
def export_company_debts_excel(request, company_id):
    """
    Export all doctor debts from a company to Excel file
    """
    company = get_object_or_404(Company, id=company_id)
    
    if not company.db_name:
        messages.error(request, 'Bu şirkətin verilənlər bazası yoxdur.')
        return redirect('master_admin:company_detail', company_id=company_id)
    
    from subscription.db_router import set_tenant_db, clear_tenant_db
    from doctors.models import Doctor
    
    try:
        # Set tenant database context
        set_tenant_db(company.db_name)
        
        # Get all doctors with debts
        doctors = Doctor.objects.select_related(
            'region', 'city', 'clinic', 'ixtisas'
        ).all().order_by('-yekun_borc')
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Borclar"
        
        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='DC2626', end_color='DC2626', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        cell_font = Font(name='Arial', size=11)
        cell_alignment = Alignment(horizontal='left', vertical='center')
        number_alignment = Alignment(horizontal='right', vertical='center')
        
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Define column headers
        headers = [
            ('Kod', 12),
            ('Ad Soyad', 30),
            ('İxtisas', 20),
            ('Bölgə', 15),
            ('Klinika', 30),
            ('Telefon', 18),
            ('Əvvəlki Borc', 15),
            ('Hesablanmış', 15),
            ('Silinən', 15),
            ('Yekun Borc', 18),
            ('Status', 12),
        ]
        
        # Write headers
        for col_num, (header, width) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
            ws.column_dimensions[cell.column_letter].width = width
        
        # Write data
        for row_num, doctor in enumerate(doctors, 2):
            ws.cell(row=row_num, column=1, value=doctor.code).font = cell_font
            ws.cell(row=row_num, column=1).alignment = cell_alignment
            ws.cell(row=row_num, column=1).border = border
            
            ws.cell(row=row_num, column=2, value=doctor.ad).font = cell_font
            ws.cell(row=row_num, column=2).alignment = cell_alignment
            ws.cell(row=row_num, column=2).border = border
            
            ws.cell(row=row_num, column=3, value=doctor.ixtisas.name if doctor.ixtisas else '-').font = cell_font
            ws.cell(row=row_num, column=3).alignment = cell_alignment
            ws.cell(row=row_num, column=3).border = border
            
            ws.cell(row=row_num, column=4, value=doctor.region.name if doctor.region else '-').font = cell_font
            ws.cell(row=row_num, column=4).alignment = cell_alignment
            ws.cell(row=row_num, column=4).border = border
            
            ws.cell(row=row_num, column=5, value=doctor.clinic.name if doctor.clinic else '-').font = cell_font
            ws.cell(row=row_num, column=5).alignment = cell_alignment
            ws.cell(row=row_num, column=5).border = border
            
            ws.cell(row=row_num, column=6, value=doctor.telefon or '-').font = cell_font
            ws.cell(row=row_num, column=6).alignment = cell_alignment
            ws.cell(row=row_num, column=6).border = border
            
            ws.cell(row=row_num, column=7, value=float(doctor.evvelki_borc or 0)).font = cell_font
            ws.cell(row=row_num, column=7).alignment = number_alignment
            ws.cell(row=row_num, column=7).border = border
            ws.cell(row=row_num, column=7).number_format = '#,##0.00'
            
            ws.cell(row=row_num, column=8, value=float(doctor.hesablanmish_miqdar or 0)).font = cell_font
            ws.cell(row=row_num, column=8).alignment = number_alignment
            ws.cell(row=row_num, column=8).border = border
            ws.cell(row=row_num, column=8).number_format = '#,##0.00'
            
            ws.cell(row=row_num, column=9, value=float(doctor.silinen_miqdar or 0)).font = cell_font
            ws.cell(row=row_num, column=9).alignment = number_alignment
            ws.cell(row=row_num, column=9).border = border
            ws.cell(row=row_num, column=9).number_format = '#,##0.00'
            
            # Yekun Borc with color coding
            debt_cell = ws.cell(row=row_num, column=10, value=float(doctor.yekun_borc or 0))
            debt_cell.font = Font(name='Arial', size=11, bold=True, color='DC2626' if doctor.yekun_borc > 0 else ('22C55E' if doctor.yekun_borc < 0 else '000000'))
            debt_cell.alignment = number_alignment
            debt_cell.border = border
            debt_cell.number_format = '#,##0.00'
            
            # Status
            status_cell = ws.cell(row=row_num, column=11, value='Aktiv' if doctor.is_active else 'Deaktiv')
            status_cell.font = cell_font
            status_cell.alignment = cell_alignment
            status_cell.border = border
            status_cell.fill = PatternFill(start_color='DCfCE7' if doctor.is_active else 'FEE2E2', end_color='DCfCE7' if doctor.is_active else 'FEE2E2', fill_type='solid')
        
        # Add summary row
        total_row = len(doctors) + 3
        ws.cell(row=total_row, column=2, value='ÜMUMİ:').font = Font(name='Arial', size=12, bold=True)
        ws.cell(row=total_row, column=2).alignment = Alignment(horizontal='right', vertical='center')
        
        total_debt = sum(float(d.yekun_borc or 0) for d in doctors)
        total_cell = ws.cell(row=total_row, column=10, value=total_debt)
        total_cell.font = Font(name='Arial', size=12, bold=True, color='DC2626' if total_debt > 0 else '000000')
        total_cell.alignment = number_alignment
        total_cell.border = border
        total_cell.number_format = '#,##0.00'
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Generate filename
        from datetime import datetime
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{company.name}_Borclar_{timestamp}.xlsx"
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        
        return response
        
    except Exception as e:
        messages.error(request, f'Excel faylı yaradılarkən xəta baş verdi: {str(e)}')
        return redirect('master_admin:company_detail', company_id=company_id)
    finally:
        clear_tenant_db()


@superuser_required
@require_http_methods(["GET", "POST"])
def import_company_doctors_full_excel(request, company_id):
    """
    Cədvəldən tam import: Bölgə, Həkim adı, İxtisas, Dərəcə, Kategoriya, Müəssisə, Borcu.
    Sistemdə yoxdursa əlavə edir (bölgə, ixtisas, müəssisə, həkim); varsa əlavə etmir.
    Sütunlar: B=Bölgə, C=Həkim adı, D=İxtisas, E=Dərəcə, F=Kategoriya, G=Müəssisə, I=Borcu.
    """
    company = get_object_or_404(Company, id=company_id)
    if not company.db_name:
        messages.error(request, "Bu şirkətin verilənlər bazası yoxdur.")
        return redirect("master_admin:company_detail", company_id=company_id)

    if request.method == "GET":
        return render(
            request,
            "master_admin/import_excel.html",
            {
                "company": company,
                "import_type": "doctors_full",
                "title": "Həkimlər cədvəlini Excel-dən Import",
                "description": "Bölgə, həkim adları, ixtisas, dərəcə, kategoriya, müəssisə və borcu olan cədvəli yükləyin. Sistemdə olmayanlar əlavə olunacaq.",
            },
        )

    upload = request.FILES.get("file")
    if not upload:
        messages.error(request, "Zəhmət olmasa Excel faylı seçin.")
        return redirect("master_admin:import_company_doctors_full", company_id=company_id)
    if not (upload.name.endswith(".xlsx") or upload.name.endswith(".xlsm")):
        messages.error(request, "Yalnız .xlsx və ya .xlsm formatlı Excel faylları dəstəklənir.")
        return redirect("master_admin:import_company_doctors_full", company_id=company_id)

    from subscription.db_router import set_tenant_db, clear_tenant_db
    from doctors.models import Doctor
    from regions.models import Region, City, Clinic, Specialization
    from decimal import Decimal, InvalidOperation

    # Sütunlar: B=2, C=3, D=4, E=5, F=6, G=7, I=9
    COL_BOLGE = 2
    COL_HEKIM = 3
    COL_IXTISAS = 4
    COL_DERECE = 5
    COL_KATEQORIYA = 6
    COL_MUESSISE = 7
    COL_BORCU = 9

    def to_decimal(val):
        if val is None or val == "":
            return Decimal("0")
        try:
            return Decimal(str(val))
        except (InvalidOperation, ValueError):
            return Decimal("0")

    def norm_degree(val):
        if not val:
            return "I"
        s = str(val).strip().upper()
        if "VIP" in s:
            return "VIP"
        if "III" in s:
            return "III"
        if "II" in s:
            return "II"
        if "I" in s:
            return "I"
        return "I"

    def norm_category(val):
        if not val:
            return "A"
        s = str(val).strip().upper().replace(" ", "")
        for code in ["A*", "A", "B", "C"]:
            if code in s or s == code:
                return code
        return "A"

    regions_created = 0
    specializations_created = 0
    clinics_created = 0
    doctors_created = 0
    skipped = 0

    try:
        wb = load_workbook(upload, data_only=True)
        ws = wb.active
        set_tenant_db(company.db_name)
        # Başlıq 1-ci və ya 2-ci sətirdə ola bilər; məlumat 2-ci sətirdən başlayır
        start_row = 2
        for row in range(start_row, ws.max_row + 1):
            bolge_name = ws.cell(row=row, column=COL_BOLGE).value
            hekim_ad = ws.cell(row=row, column=COL_HEKIM).value
            if not bolge_name or not hekim_ad:
                skipped += 1
                continue
            bolge_name = str(bolge_name).strip()
            hekim_ad = str(hekim_ad).strip()
            # Başlıq sətirini atla
            if hekim_ad.lower() in ("həkim adı", "hekim adi", "həkim adi") or bolge_name.lower() == "bölgə":
                skipped += 1
                continue
            if not bolge_name or not hekim_ad:
                skipped += 1
                continue

            # Bölgə — yoxdursa əlavə et
            region = Region.objects.filter(name__iexact=bolge_name).first()
            if not region:
                region = Region.objects.create(name=bolge_name)
                regions_created += 1

            # Şəhər (klinika üçün lazımdır) — bölgə üçün default bir şəhər
            default_city = City.objects.filter(region=region).first()
            if not default_city:
                default_city = City.objects.create(region=region, name=region.name)

            # İxtisas — yoxdursa əlavə et
            ixtisas_name = ws.cell(row=row, column=COL_IXTISAS).value
            ixtisas_name = str(ixtisas_name).strip() if ixtisas_name else ""
            if not ixtisas_name:
                ixtisas_name = "—"
            ixtisas = Specialization.objects.filter(name__iexact=ixtisas_name).first()
            if not ixtisas:
                ixtisas = Specialization.objects.create(name=ixtisas_name)
                specializations_created += 1

            # Müəssisə (Klinika) — yoxdursa əlavə et
            muessise = ws.cell(row=row, column=COL_MUESSISE).value
            muessise = str(muessise).strip() if muessise else ""
            clinic = None
            if muessise:
                clinic = Clinic.objects.filter(region=region, name__iexact=muessise).first()
                if not clinic:
                    clinic = Clinic.objects.create(
                        name=muessise,
                        region=region,
                        city=default_city,
                        address="",
                    )
                    clinics_created += 1

            # Həkim — yalnız yoxdursa əlavə et (ad + bölgə ilə yoxla)
            existing = Doctor.objects.filter(ad__iexact=hekim_ad, region=region).first()
            if existing:
                skipped += 1
                continue

            degree_code = norm_degree(ws.cell(row=row, column=COL_DERECE).value)
            category_code = norm_category(ws.cell(row=row, column=COL_KATEQORIYA).value)
            borc = to_decimal(ws.cell(row=row, column=COL_BORCU).value)

            Doctor.objects.create(
                ad=hekim_ad,
                region=region,
                city=default_city,
                clinic=clinic,
                ixtisas=ixtisas,
                degree=degree_code,
                category=category_code,
                telefon="-",
                evvelki_borc=borc,
                hesablanmish_miqdar=Decimal("0"),
                silinen_miqdar=Decimal("0"),
            )
            doctors_created += 1

        messages.success(
            request,
            f"Import tamamlandı: {regions_created} bölgə, {specializations_created} ixtisas, "
            f"{clinics_created} klinika, {doctors_created} həkim əlavə olundu; {skipped} sətir (artıq mövcud) ötürüldü.",
        )
        return redirect("master_admin:company_detail", company_id=company_id)
    except Exception as e:
        messages.error(request, f"Import xətası: {str(e)}")
        return redirect("master_admin:import_company_doctors_full", company_id=company_id)
    finally:
        clear_tenant_db()


@superuser_required
@require_http_methods(["POST"])
def zero_company_doctors_debts(request, company_id):
    """Bütün həkimlərin yalnız Əvvəlki borcunu 0 et (yekun borc yenidən hesablanacaq)."""
    company = get_object_or_404(Company, id=company_id)
    if not company.db_name:
        messages.error(request, "Bu şirkətin verilənlər bazası yoxdur.")
        return redirect("master_admin:company_detail", company_id=company_id)

    from subscription.db_router import set_tenant_db, clear_tenant_db
    from doctors.models import Doctor
    from decimal import Decimal

    try:
        set_tenant_db(company.db_name)
        doctors = Doctor.objects.all()
        count = 0
        for doctor in doctors:
            doctor.evvelki_borc = Decimal("0")
            doctor.save()
            count += 1
        messages.success(request, f"Həkimlərin əvvəlki borcları sıfırlandı. Cəmi {count} həkim yeniləndi.")
    except Exception as e:
        messages.error(request, f"Borclar sıfırlanarkən xəta: {str(e)}")
    finally:
        clear_tenant_db()
    return redirect("master_admin:company_detail", company_id=company_id)


@superuser_required
@require_http_methods(["GET", "POST"])
def import_company_debts_excel(request, company_id):
    """
    Import/update doctor debts from Excel file for a specific company.
    
    Gözlənilən sütunlar: Bölgə, Həkim adı, Yekun borc
    """
    company = get_object_or_404(Company, id=company_id)

    if not company.db_name:
        messages.error(request, "Bu şirkətin verilənlər bazası yoxdur.")
        return redirect("master_admin:company_detail", company_id=company_id)

    if request.method == "GET":
        return render(
            request,
            "master_admin/import_excel.html",
            {
                "company": company,
                "import_type": "debts",
                "title": "Borcları Excel-dən Import",
                "description": "Bölgə, Həkim adı və Yekun borc sütunlarını ehtiva edən Excel faylını yükləyin.",
            },
        )

    # POST: handle upload
    upload = request.FILES.get("file")
    if not upload:
        messages.error(request, "Zəhmət olmasa Excel faylı seçin.")
        return redirect("master_admin:import_company_debts", company_id=company_id)

    if not (upload.name.endswith(".xlsx") or upload.name.endswith(".xlsm")):
        messages.error(request, "Yalnız .xlsx və ya .xlsm formatlı Excel faylları dəstəklənir.")
        return redirect("master_admin:import_company_debts", company_id=company_id)

    from subscription.db_router import set_tenant_db, clear_tenant_db
    from doctors.models import Doctor
    from regions.models import Region
    from decimal import Decimal, InvalidOperation

    updated = 0
    skipped = 0

    try:
        wb = load_workbook(upload, data_only=True)
        ws = wb.active

        # A = Bölgə, B = Həkim adı, C = Yekun borc (sütun mövqeyinə görə)
        col_bolge = 1   # A
        col_hekim = 2   # B
        col_yekun = 3   # C
        header_row = 1  # Birinci sətir başlıq ola bilər, məlumat 2-ci sətirdən

        set_tenant_db(company.db_name)

        def to_decimal(val):
            if val is None or val == "":
                return None
            try:
                return Decimal(str(val))
            except (InvalidOperation, ValueError):
                return None

        for row in range(header_row + 1, ws.max_row + 1):
            bolge_name = ws.cell(row=row, column=col_bolge).value
            hekim_ad = ws.cell(row=row, column=col_hekim).value
            yekun_val = to_decimal(ws.cell(row=row, column=col_yekun).value)

            if not bolge_name or not hekim_ad:
                skipped += 1
                continue
            bolge_name = str(bolge_name).strip()
            hekim_ad = str(hekim_ad).strip()
            if not bolge_name or not hekim_ad:
                skipped += 1
                continue
            if yekun_val is None:
                yekun_val = Decimal("0")

            region = Region.objects.filter(name__iexact=bolge_name).first()
            if not region:
                skipped += 1
                continue

            doctor = Doctor.objects.filter(ad__iexact=hekim_ad, region=region).first()
            if not doctor:
                skipped += 1
                continue

            doctor.evvelki_borc = yekun_val
            doctor.hesablanmish_miqdar = Decimal("0")
            doctor.silinen_miqdar = Decimal("0")
            doctor.save()
            updated += 1

        messages.success(
            request,
            f"Borcların importu tamamlandı. Yenilənən həkim sayı: {updated}, ötürülən sətir: {skipped}.",
        )
        return redirect("master_admin:company_detail", company_id=company_id)

    except Exception as e:
        messages.error(request, f"Excel faylı oxunarkən xəta baş verdi: {str(e)}")
        return redirect("master_admin:import_company_debts", company_id=company_id)
    finally:
        clear_tenant_db()


@superuser_required
@require_http_methods(["GET", "POST"])
def import_company_drugs_excel(request, company_id):
    """
    Import/add/update drugs from Excel file for a specific company.
    
    Expected columns (header row, in any order):
    - Ad        -> ad (required)
    - Tam Ad    -> tam_ad (optional)
    - Qiymət    -> qiymet (required)
    - Komissiya -> komissiya (optional, default 0)
    """
    company = get_object_or_404(Company, id=company_id)

    if not company.db_name:
        messages.error(request, "Bu şirkətin verilənlər bazası yoxdur.")
        return redirect("master_admin:company_detail", company_id=company_id)

    if request.method == "GET":
        return render(
            request,
            "master_admin/import_excel.html",
            {
                "company": company,
                "import_type": "drugs",
                "title": "Dərmanları Excel-dən Import",
                "description": "Dərman siyahısını ehtiva edən Excel faylını yükləyin.",
            },
        )

    upload = request.FILES.get("file")
    if not upload:
        messages.error(request, "Zəhmət olmasa Excel faylı seçin.")
        return redirect("master_admin:import_company_drugs", company_id=company_id)

    if not (upload.name.endswith(".xlsx") or upload.name.endswith(".xlsm")):
        messages.error(request, "Yalnız .xlsx və ya .xlsm formatlı Excel faylları dəstəklənir.")
        return redirect("master_admin:import_company_drugs", company_id=company_id)

    from subscription.db_router import set_tenant_db, clear_tenant_db
    from drugs.models import Drug
    from decimal import Decimal, InvalidOperation

    created = 0
    updated = 0
    skipped = 0

    try:
        wb = load_workbook(upload, data_only=True)
        ws = wb.active

        # Build header map
        header_row = 1
        headers = {}
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=header_row, column=col).value
            if not value:
                continue
            header_text = str(value).strip().lower()
            headers[header_text] = col

        # Required columns
        if "ad" not in headers or "qiymət" not in headers:
            messages.error(
                request,
                'Sütun tapılmadı: "Ad" və "Qiymət" mütləqdir. Zəhmət olmasa şablona uyğun fayl yükləyin.',
            )
            return redirect("master_admin:import_company_drugs", company_id=company_id)

        col_ad = headers["ad"]
        col_tam_ad = headers.get("tam ad")
        col_price = headers["qiymət"]
        col_comm = headers.get("komissiya")

        set_tenant_db(company.db_name)

        def to_decimal(val):
            if val is None or val == "":
                return None
            try:
                return Decimal(str(val))
            except (InvalidOperation, ValueError):
                return None

        for row in range(header_row + 1, ws.max_row + 1):
            name = ws.cell(row=row, column=col_ad).value
            if not name:
                skipped += 1
                continue
            name = str(name).strip()
            if not name:
                skipped += 1
                continue

            price = to_decimal(ws.cell(row=row, column=col_price).value)
            if price is None:
                skipped += 1
                continue

            tam_ad = (
                str(ws.cell(row=row, column=col_tam_ad).value).strip()
                if col_tam_ad
                and ws.cell(row=row, column=col_tam_ad).value is not None
                else ""
            )

            komissiya = (
                to_decimal(ws.cell(row=row, column=col_comm).value)
                if col_comm
                else None
            )
            if komissiya is None:
                komissiya = Decimal("0")

            # Mövcud dərmanı ad üzrə tap
            drug = Drug.objects.filter(ad=name).first()

            if drug:
                # Yenilə
                drug.ad = name
                drug.tam_ad = tam_ad or name
                drug.qiymet = price
                drug.komissiya = komissiya
                drug.is_active = True
                drug.save()
                updated += 1
            else:
                # Yeni yarat (buraxılış forması, dozaj, barkod default/boş)
                drug = Drug.objects.create(
                    ad=name,
                    tam_ad=tam_ad or name,
                    qiymet=price,
                    komissiya=komissiya,
                    is_active=True,
                )
                created += 1

        messages.success(
            request,
            f"Dərmanların importu tamamlandı. Yaradılan: {created}, yenilənən: {updated}, ötürülən: {skipped}.",
        )
        return redirect("master_admin:company_detail", company_id=company_id)

    except Exception as e:
        messages.error(request, f"Excel faylı oxunarkən xəta baş verdi: {str(e)}")
        return redirect("master_admin:import_company_drugs", company_id=company_id)
    finally:
        clear_tenant_db()

