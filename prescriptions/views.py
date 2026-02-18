from collections import OrderedDict
from datetime import datetime
from django.template.loader import render_to_string
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from subscription.decorators import subscription_required
from doctors.models import Doctor, DoctorPayment
from drugs.models import Drug
from regions.models import Region 
from .models import Prescription, PrescriptionItem
from reports.models import MonthlyDoctorReport
from datetime import date  


@login_required
@subscription_required
def add_prescription(request):
    """Add new prescription/recipe"""
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Get form data
                region_id = request.POST.get('region_id')
                doctor_id = request.POST.get('doctor_id')
                date = request.POST.get('date')
           
                # Validation
                if not region_id:
                    messages.error(request, 'Bölgə seçilməlidir.')
                    return redirect('prescriptions:add')
                
                if not doctor_id or not date:
                    messages.error(request, 'Həkim və Tarix doldurulmalıdır.')
                    return redirect('prescriptions:add')
                
                # Collect drug quantities
                selected_drugs = []
                for key, value in request.POST.items():
                    if key.startswith('drug_'):
                        try:
                            drug_id = int(key.split('_')[1])
                            quantity = int(value) if value else 0
                        except (ValueError, IndexError):
                            continue
                        
                        if quantity > 0:
                            selected_drugs.append((drug_id, quantity))
                
                if not selected_drugs:
                    messages.error(request, 'Ən azı bir dərman üçün miqdar daxil edilməlidir.')
                    return redirect('prescriptions:add')
                
                # Get doctor
                doctor = Doctor.objects.get(id=doctor_id)
                region = Region.objects.get(id=region_id)
                
                # Validate date against last closed report
                prescription_date = datetime.strptime(date, '%Y-%m-%d').date()
                last_report = MonthlyDoctorReport.objects.filter(
                    region_id=region_id
                ).order_by('-year', '-month').first()
                
                if last_report:
                    # Calculate the next month after last closed report
                    last_closed_date = date(last_report.year, last_report.month, 1)
                    # Get first day of next month
                    if last_report.month == 12:
                        next_month_date = date(last_report.year + 1, 1, 1)
                    else:
                        next_month_date = date(last_report.year, last_report.month + 1, 1)
                    
                    # Check if prescription date is before the allowed month
                    prescription_month_start = date(prescription_date.year, prescription_date.month, 1)
                    
                    if prescription_month_start < next_month_date:
                        messages.error(
                            request, 
                            f'Bu bölgə üçün ən son bağlanan hesabat {last_report.month:02d}/{last_report.year} ayıdır. '
                            f'Qeydiyyat yalnız {next_month_date.month:02d}/{next_month_date.year} ayından başlayaraq əlavə edilə bilər.'
                        )
                        return redirect('prescriptions:add')
                
                # Create prescription
                prescription = Prescription.objects.create(
                    region=region,
                    doctor=doctor,
                    date=date
                )
                
                # Add drugs to prescription
                drug_ids = [drug_id for drug_id, _ in selected_drugs]
                drugs_map = {drug.id: drug for drug in Drug.objects.filter(id__in=drug_ids)}
                
                for drug_id, quantity in selected_drugs:
                    drug = drugs_map.get(drug_id)
                    if not drug:
                        continue
                    PrescriptionItem.objects.create(
                        prescription=prescription,
                        drug=drug,
                        quantity=quantity,
                        unit_price=drug.qiymet
                    )
                
                messages.success(request, f'Resept uğurla əlavə edildi! ({prescription.drug_count} dərman)')
                return redirect('prescriptions:add')
                
        except Region.DoesNotExist:
            messages.error(request, 'Bölgə tapılmadı.')
        except Doctor.DoesNotExist:
            messages.error(request, 'Həkim tapılmadı.')
        except Drug.DoesNotExist:
            messages.error(request, 'Dərman tapılmadı.')
        except Exception as e:
            messages.error(request, f'Xəta baş verdi: {str(e)}')
        
        return redirect('prescriptions:add')
    
    # GET request - show form
    drugs = Drug.objects.filter(is_active=True).order_by('ad')
    regions = Region.objects.all().order_by('name')
    
    # Safely load recent prescriptions without accessing problematic Decimal fields
    try:
        recent_prescriptions_qs = Prescription.objects.select_related(
            'region', 'doctor'
        ).order_by('-date')[:5]
        
        # Convert to list of dicts to avoid accessing PrescriptionItem with invalid Decimal data
        recent_prescriptions_list = []
        for prescription in recent_prescriptions_qs:
            try:
                # Safely get items count and basic info without loading Decimal fields
                from django.db.models import Count
                item_count = PrescriptionItem.objects.filter(
                    prescription_id=prescription.id
                ).count()
                
                # Get items with only safe fields
                items_list = []
                try:
                    items_qs = PrescriptionItem.objects.filter(
                        prescription_id=prescription.id
                    ).select_related('drug').only('id', 'quantity', 'drug__ad', 'drug__id')
                    
                    for item in items_qs:
                        try:
                            items_list.append({
                                'drug': {'ad': item.drug.ad if item.drug else '-'},
                                'quantity': item.quantity or 0,
                            })
                        except Exception:
                            continue
                except Exception:
                    pass
                
                recent_prescriptions_list.append({
                    'id': prescription.id,
                    'date': prescription.date,
                    'region': prescription.region,
                    'doctor': prescription.doctor,
                    'patient_name': getattr(prescription, 'patient_name', None),
                    'items': items_list,
                    'item_count': item_count,
                })
            except Exception:
                continue
        recent_prescriptions = recent_prescriptions_list
    except Exception:
        recent_prescriptions = []
    
    context = {
        'regions': regions,
        'drugs': drugs,
        'recent_prescriptions': recent_prescriptions,
    }
    
    return render(request, 'prescriptions/add.html', context)


@login_required
@subscription_required
def prescription_list(request):
    """List all prescriptions with filterable monthly summary"""

    filters = _build_filters(request)
    doctor_summary, stats = _get_filtered_summary(filters, request)
    regions = _get_regions()
    drugs = _get_drugs()
    region_lookup = {str(region.id): region for region in regions}

    context = {
        'doctor_summary': doctor_summary,
        'drugs': drugs,
        'regions': regions,
        'region_lookup': region_lookup,
        'filters': filters,
        'stats': stats,
        'selected_region_label': _get_selected_region_label(filters, region_lookup),
        'region_selected': bool(filters['region']),
    }

    return render(request, 'prescriptions/list.html', context)


@login_required
@subscription_required
def prescription_monthly_report(request):
    from django.db.models import Sum

    # Filter by month if provided
    month = request.GET.get('month')
    year = request.GET.get('year')
    drugs = Drug.objects.filter(is_active=True).order_by('ad')


    if not month:
        month = timezone.now().month
    if not year:
        year = timezone.now().year

    # Filter prescriptions by date
    prescriptions = Prescription.objects.filter(
        date__month=month,
        date__year=year
    ).select_related('doctor').prefetch_related('items__drug')

    # Prepare doctor-based summary
    doctor_summary = {}

    for p in prescriptions:
        doc = p.doctor

        if doc.id not in doctor_summary:
            doctor_summary[doc.id] = {
                'doctor': doc,
                'drugs': {},
                'total_amount': 0
            }

        for item in p.items.all():
            drug_name = item.drug.ad

            # Count quantity (sum)
            doctor_summary[doc.id]['drugs'][drug_name] = (
                doctor_summary[doc.id]['drugs'].get(drug_name, 0) + item.quantity
            )

            # Total price
            doctor_summary[doc.id]['total_amount'] += item.quantity * item.unit_price

    context = {
        'month': month,
        'year': year,
        'doctor_summary': doctor_summary.values(),
        'drugs': drugs,
    }

    return render(request, 'prescriptions/list.html', context)



@login_required
@subscription_required
def doctors_by_region(request, region_id):
    """Return doctors for a given region"""
    doctors = Doctor.objects.filter(region_id=region_id, is_active=True).order_by('ad')
    data = [
        {
            'id': doctor.id,
            'name': doctor.ad,
            'code': doctor.code
        }
        for doctor in doctors
    ]
    return JsonResponse({'doctors': data})


@login_required
@subscription_required
def get_last_closed_report(request, region_id):
    """Get the last closed monthly report for a region"""
    try:
        # Find the most recent closed report for this region
        last_report = MonthlyDoctorReport.objects.filter(
            region_id=region_id
        ).order_by('-year', '-month').first()
        
        if last_report:
            return JsonResponse({
                'last_closed_year': last_report.year,
                'last_closed_month': last_report.month,
                'has_closed_report': True
            })
        else:
            # No closed reports for this region - allow all dates
            return JsonResponse({
                'last_closed_year': None,
                'last_closed_month': None,
                'has_closed_report': False
            })
    except Exception as e:
        return JsonResponse({
            'error': str(e),
            'has_closed_report': False
        }, status=500)

@login_required
@subscription_required
def filter_prescriptions_ajax(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Yalnız POST müraciətinə icazə verilir.'}, status=405)

    filters = _build_filters(request)
    doctor_summary, stats = _get_filtered_summary(filters, request)
    regions = _get_regions()
    region_lookup = {str(region.id): region for region in regions}

    context = {
        'doctor_summary': doctor_summary,
        'stats': stats,
        'drugs': _get_drugs(),
        'selected_region_label': _get_selected_region_label(filters, region_lookup),
        'region_selected': bool(filters['region']),
    }

    html = render_to_string("prescriptions/table_partial.html", context, request=request)
    return JsonResponse({"html": html})


def _build_filters(request):
    data = request.POST if request.method == 'POST' else request.GET
    return {
        'region': (data.get('region') or '').strip(),
        'start_date': (data.get('start_date') or '').strip(),
        'end_date': (data.get('end_date') or '').strip(),
        'doctor': (data.get('doctor') or '').strip(),
    }


def _parse_date(value, request=None):
    if not value:
        return None
    try:
        return datetime.strptime(value, '%Y-%m-%d').date()
    except ValueError:
        if request:
            messages.error(request, 'Tarix formatı yanlışdır. YYYY-MM-DD olaraq daxil edin.')
        return None


def _get_filtered_summary(filters, request=None):
    parse_request = request if request is not None and request.method == 'GET' else None
    start_date = _parse_date(filters['start_date'], request=parse_request)
    end_date = _parse_date(filters['end_date'], request=parse_request)

    if not filters['region']:
        return [], {'doctor_count': 0, 'total_amount': 0, 'total_quantity': 0}

    doctor_qs = (
        Doctor.objects
        .select_related('region')
        .filter(region_id=filters['region'])
    )

    if filters['doctor']:
        doctor_qs = doctor_qs.filter(
            Q(ad__icontains=filters['doctor']) |
            Q(code__icontains=filters['doctor'])
        )

    doctor_qs = doctor_qs.order_by('ad')

    doctor_summary = OrderedDict()

    for doctor in doctor_qs:
        doctor_summary[doctor.id] = {
            'doctor': doctor,
            'region': doctor.region,
            'drugs': {},
            'total_amount': 0,
            'total_quantity': 0,
            'last_date': None,
        }

    if not doctor_summary:
        return [], {'doctor_count': 0, 'total_amount': 0, 'total_quantity': 0}

    doctor_ids = list(doctor_summary.keys())

    prescriptions = (
        Prescription.objects
        .select_related('doctor', 'region')
        .prefetch_related('items__drug')
        .filter(doctor_id__in=doctor_ids)
        .order_by('-date')
    )

    if start_date:
        prescriptions = prescriptions.filter(date__gte=start_date)

    if end_date:
        prescriptions = prescriptions.filter(date__lte=end_date)

    for prescription in prescriptions:
        summary = doctor_summary.get(prescription.doctor_id)
        if not summary:
            continue

        if prescription.date and (not summary['last_date'] or prescription.date > summary['last_date']):
            summary['last_date'] = prescription.date

        for item in prescription.items.all():
            drug_name = item.drug.ad
            summary['drugs'][drug_name] = summary['drugs'].get(drug_name, 0) + item.quantity
            summary['total_quantity'] += item.quantity
            summary['total_amount'] += item.quantity * item.unit_price

    summary_list = list(doctor_summary.values())
    _attach_last_payments(summary_list)

    stats = {
        'doctor_count': len(summary_list),
        'total_amount': sum(entry['total_amount'] for entry in summary_list),
        'total_quantity': sum(entry['total_quantity'] for entry in summary_list),
    }

    return summary_list, stats


def _attach_last_payments(summary_list, limit=1):
    if not summary_list:
        return

    doctor_ids = [entry['doctor'].id for entry in summary_list if entry.get('doctor')]
    if not doctor_ids:
        return

    payments_map = {doctor_id: [] for doctor_id in doctor_ids}
    payments = (
        DoctorPayment.objects
        .select_related('doctor')
        .filter(doctor_id__in=doctor_ids)
        .order_by('doctor_id', '-date', '-created_at')
    )

    for payment in payments:
        bucket = payments_map.get(payment.doctor_id)
        if bucket is None:
            continue
        if len(bucket) < limit:
            bucket.append(payment)

    for entry in summary_list:
        entry['last_payments'] = payments_map.get(entry['doctor'].id, [])


def _get_regions():
    return Region.objects.order_by('name')


def _get_drugs():
    return Drug.objects.filter(is_active=True).order_by('ad')


def _get_selected_region_label(filters, region_lookup):
    selected = region_lookup.get(filters['region'])
    return selected.name if selected else ''


def _get_last_payments(doctor):
    try:
        return DoctorPayment.objects.filter(doctor=doctor).order_by('-date')[:3].values('date', 'amount')
        
    except Exception as e:
        return []
    return []


@login_required
@subscription_required
def export_prescriptions_excel(request):
    """
    Export prescriptions to Excel file with formatting
    """
    # Get filters from request
    filters = _build_filters(request)
    
    # Get filtered prescriptions
    prescriptions = Prescription.objects.select_related(
        'doctor', 'region'
    ).prefetch_related('items__drug').all()
    
    # Apply filters
    if filters['region']:
        prescriptions = prescriptions.filter(region_id=filters['region'])
    
    if filters['start_date']:
        try:
            start_date = datetime.strptime(filters['start_date'], '%Y-%m-%d').date()
            prescriptions = prescriptions.filter(date__gte=start_date)
        except ValueError:
            pass
    
    if filters['end_date']:
        try:
            end_date = datetime.strptime(filters['end_date'], '%Y-%m-%d').date()
            prescriptions = prescriptions.filter(date__lte=end_date)
        except ValueError:
            pass
    
    if filters['doctor']:
        prescriptions = prescriptions.filter(
            Q(doctor__ad__icontains=filters['doctor']) |
            Q(doctor__code__icontains=filters['doctor'])
        )
    
    prescriptions = prescriptions.order_by('-date', '-created_at')
    
    # Get ALL active drugs (not just from prescriptions) - so new drugs automatically get columns
    all_drugs = Drug.objects.filter(is_active=True).order_by('ad')
    drugs_list = [(drug.id, drug.ad) for drug in all_drugs]
    drug_columns = {drug_id: idx for idx, (drug_id, _) in enumerate(drugs_list, start=5)}
    
    # Get last payments for all doctors
    doctor_ids = list(set(prescriptions.values_list('doctor_id', flat=True)))
    payments_map = {}
    if doctor_ids:
        from doctors.models import DoctorPayment
        # Get the most recent payment for each doctor
        for doctor_id in doctor_ids:
            last_payment = (
                DoctorPayment.objects
                .filter(doctor_id=doctor_id)
                .order_by('-date', '-created_at')
                .first()
            )
            if last_payment:
                payments_map[doctor_id] = last_payment
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Resept Qeydiyyatları"
    
    # Define styles
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    cell_font = Font(name='Arial', size=11)
    cell_alignment = Alignment(horizontal='left', vertical='center')
    number_alignment = Alignment(horizontal='right', vertical='center')
    date_alignment = Alignment(horizontal='center', vertical='center')
    
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Define fixed column headers
    fixed_headers = [
        ('№', 6),
        ('Bölgə', 18),
        ('Son Ödəniş', 15),
        ('Kod', 12),
    ]
    
    # Add drug columns
    drug_headers = [(drug_name, 15) for _, drug_name in drugs_list]
    
    # Add Cem column at the end
    all_headers = fixed_headers + drug_headers + [('Cem', 12)]
    
    # Write headers
    for col_num, (header, width) in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Write data - one row per prescription
    row_num = 2
    prescription_num = 1
    
    for prescription in prescriptions:
        # Get all items for this prescription
        items = prescription.items.all()
        
        if not items:
            continue  # Skip prescriptions without items
        
        # Get last payment for this doctor
        last_payment = payments_map.get(prescription.doctor_id)
        last_payment_str = '-'
        if last_payment:
            last_payment_str = f"{last_payment.date.strftime('%d.%m.%Y')} {float(last_payment.amount):.2f} ₼"
        
        # Create a dictionary for drug quantities in this prescription
        prescription_drugs = {}
        total_quantity = 0
        for item in items:
            if item.drug:
                drug_id = item.drug.id
                quantity = item.quantity
                prescription_drugs[drug_id] = quantity
                total_quantity += quantity
        
        # Write one row per prescription
        # № (Column 1)
        cell = ws.cell(row=row_num, column=1)
        cell.value = prescription_num
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = border
        
        # Bölgə (Column 2)
        cell = ws.cell(row=row_num, column=2)
        cell.value = prescription.region.name if prescription.region else '-'
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = border
        
        # Son Ödəniş (Column 3)
        cell = ws.cell(row=row_num, column=3)
        cell.value = last_payment_str
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = border
        
        # Kod (Column 4)
        cell = ws.cell(row=row_num, column=4)
        cell.value = prescription.doctor.code if prescription.doctor else '-'
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = border
        
        # Write drug quantities in their respective columns
        for drug_id, col_num in drug_columns.items():
            quantity = prescription_drugs.get(drug_id, 0)
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = quantity if quantity > 0 else ''
            cell.font = cell_font
            cell.alignment = number_alignment
            cell.border = border
        
        # Cem (Last column) - total quantity
        cem_col = len(all_headers)
        cell = ws.cell(row=row_num, column=cem_col)
        cell.value = total_quantity
        cell.font = cell_font
        cell.alignment = number_alignment
        cell.border = border
        
        row_num += 1
        prescription_num += 1
    
    # Freeze first row
    ws.freeze_panes = 'A2'
    
    # Create HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'Resept_Qeydiyyatlari_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response



