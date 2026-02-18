# reports/views.py
from collections import OrderedDict
from datetime import date
from decimal import Decimal

from django.db.models import Q, Prefetch
from django.http import HttpResponse
from django.shortcuts import render

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from doctors.models import Doctor, DoctorPayment
from drugs.models import Drug
from prescriptions.models import Prescription, PrescriptionItem
from regions.models import Region
from .models import MonthlyDoctorReport
from django.contrib import messages
from django.shortcuts import redirect
from django.db import transaction


def monthly_reports(request):
    """Monthly doctor performance + financial report."""
    today = date.today()
    filters, doctor_rows, drugs = prepare_reports_data(request)

    regions = Region.objects.order_by("name")
    years = range(today.year - 2, today.year + 3)

    stats = {
        "doctor_count": len(doctor_rows),
        "total_debt": sum(row["yekun_borc"] for row in doctor_rows),
        "total_prescriptions": sum(row["total_quantity"] for row in doctor_rows),
    }

    table_info = build_table_info(filters, regions, stats)

    return render(
        request,
        "prescriptions/reports.html",
        {
            "filters": filters,
            "regions": regions,
            "drugs": drugs,
            "doctor_rows": doctor_rows,
            "years": years,
            "months": MONTH_CHOICES,
            "stats": stats,
            "table_info": table_info,
        }
    )



# -----------------------------
# HELPERS
# -----------------------------

MONTH_CHOICES = [
    ("", "Ay seçin"),
    ("1", "Yanvar"), ("2", "Fevral"), ("3", "Mart"), ("4", "Aprel"),
    ("5", "May"), ("6", "İyun"), ("7", "İyul"), ("8", "Avqust"),
    ("9", "Sentyabr"), ("10", "Oktyabr"), ("11", "Noyabr"), ("12", "Dekabr"),
]


def safe_int(value, fallback=None):
    try:
        return int(value) if value else fallback
    except (ValueError, TypeError):
        return fallback


def build_table_info(filters, regions, stats):
    info = []

    month_label = next((label for val, label in MONTH_CHOICES if val == filters["month"]), None)
    info.append(month_label or "Bütün aylar")

    info.append(filters["year"] or "")

    if filters["region"]:
        region = regions.filter(id=filters["region"]).first()
        info.append(region.name if region else "Seçilmiş bölgə")
    else:
        info.append("Bütün bölgələr")

    info.append(f"Həkim sayı: {stats['doctor_count']}")

    return " · ".join(info)


def prepare_reports_data(request):
    """
    MƏNTİQİ:
    1) Əgər seçilmiş ay üçün snapshot varsa → arxivdən oxu
    2) Snapshot yoxdursa → yoxla:
         - Əgər ay bağlanmış son aydan böyükdürsə → yeni ay → hamısı 0 çıxmalıdır
         - Əks halda → canlı hesablama et
    3) Resepti olmayan həkimlər də cədvəldə görünür → hamısı 0 ilə
    """

    today = date.today()

    # -----------------------------
    # FILTERS
    # -----------------------------
    filters = {
        "region": (request.GET.get("region") or "").strip(),
        "month": (request.GET.get("month") or "").strip(),
        "year": (request.GET.get("year") or str(today.year)).strip(),
        "doctor": (request.GET.get("doctor") or "").strip(),
    }

    month_int = safe_int(filters["month"])
    year_int = safe_int(filters["year"], today.year)

    # -----------------------------
    # REGION REQUIRED - If no region selected, return empty data
    # -----------------------------
    if not filters["region"]:
        drugs = Drug.objects.filter(is_active=True).order_by("ad")
        return filters, [], drugs

    # -----------------------------
    # LOAD DOCTORS (all doctors in region filter)
    # -----------------------------
    doctors = Doctor.objects.select_related("region", "ixtisas").filter(region_id=filters["region"])

    if filters["doctor"]:
        doctors = doctors.filter(
            Q(ad__icontains=filters["doctor"]) |
            Q(code__icontains=filters["doctor"])
        )

    doctors = list(doctors)

    # -----------------------------
    # 1) SNAPSHOT MODE?
    # -----------------------------
    snapshot_mode = False
    snapshot_qs = MonthlyDoctorReport.objects.none()

    if month_int and year_int and filters["region"]:
        snapshot_qs = MonthlyDoctorReport.objects.select_related(
            "doctor", "doctor__region", "doctor__ixtisas"
        ).filter(year=year_int, month=month_int, region_id=filters["region"])

        if filters["doctor"]:
            snapshot_qs = snapshot_qs.filter(
                Q(doctor__ad__icontains=filters["doctor"]) |
                Q(doctor__code__icontains=filters["doctor"])
            )

        snapshot_mode = snapshot_qs.exists()

    # =========================================================================
    # 1) SNAPSHOT VARSA → BURDAN OXU
    # =========================================================================
    if snapshot_mode:
        doctor_rows = []
        for report in snapshot_qs:
            doctor_rows.append({
                "doctor": report.doctor,
                "drugs": dict(report.drugs_data or {}),
                "total_quantity": report.total_quantity,
                "evvelki_borc": report.evvelki_borc,
                "hesablanan": report.hesablanan,
                "silinen_miqdar": report.silinen_miqdar,
                "datasiya": report.datasiya,
                "payments": {
                    "avans": report.avans,
                    "investisiya": report.investisiya,
                    "geriqaytarma": report.geriqaytarma,
                },
                "yekun_borc": report.yekun_borc,
            })

        drugs = Drug.objects.filter(is_active=True).order_by("ad")
        return filters, doctor_rows, drugs

    # =========================================================================
    # 2) ƏVVƏLCƏ SON BAĞLANMIŞ AYI TAP
    # =========================================================================
    last_snapshot = MonthlyDoctorReport.objects.order_by("-year", "-month").first()

    if last_snapshot:
        last_closed_year = last_snapshot.year
        last_closed_month = last_snapshot.month

        # Soruşulan ay seçilməyibsə → gələcək ay yoxlaması etmə
        if not month_int:
            is_future_month = False
        else:
            is_future_month = (
                (year_int > last_closed_year) or
                (year_int == last_closed_year and month_int > last_closed_month)
            )

        if is_future_month:
            doctor_rows = []
            for doctor in doctors:
                doctor_rows.append({
                    "doctor": doctor,
                    "drugs": {},
                    "total_quantity": 0,
                    "evvelki_borc": doctor.evvelki_borc,
                    "hesablanan": Decimal("0"),
                    "silinen_miqdar": Decimal("0"),
                    "datasiya": Decimal("0"),
                    "payments": {
                        "avans": Decimal("0"),
                        "investisiya": Decimal("0"),
                        "geriqaytarma": Decimal("0")
                    },
                    "yekun_borc": doctor.evvelki_borc,
                })

            drugs = Drug.objects.filter(is_active=True).order_by("ad")
            return filters, doctor_rows, drugs


    # =========================================================================
    # 3) SNAPSHOT YOXDUR → CANLI HESABLA
    # =========================================================================

    # doctor_summary-i əvvəlcədən 0-lar ilə doldururuq
    doctor_summary = OrderedDict()
    for doctor in doctors:
        doctor_summary[doctor.id] = {
            "doctor": doctor,
            "drugs": {},
            "total_quantity": 0,
            "evvelki_borc": doctor.evvelki_borc,
            "hesablanan": doctor.hesablanmish_miqdar,
            "silinen_miqdar": doctor.silinen_miqdar,
            "datasiya": Decimal("0"),
            "payments": {
                "avans": Decimal("0"),
                "investisiya": Decimal("0"),
                "geriqaytarma": Decimal("0")
            },
            "yekun_borc": Decimal("0"),
        }

    # Prefetch prescription data
    item_prefetch = Prefetch(
        "items",
        queryset=PrescriptionItem.objects.select_related("drug").only("drug__ad", "quantity")
    )

    prescriptions = Prescription.objects.select_related(
        "doctor", "doctor__region", "doctor__ixtisas"
    ).prefetch_related(item_prefetch)

    if filters["region"]:
        prescriptions = prescriptions.filter(
            Q(region_id=filters["region"]) |
            Q(doctor__region_id=filters["region"])
        )

    if month_int:
        prescriptions = prescriptions.filter(date__month=month_int)

    prescriptions = prescriptions.filter(date__year=year_int)

    if filters["doctor"]:
        prescriptions = prescriptions.filter(
            Q(doctor__ad__icontains=filters["doctor"]) |
            Q(doctor__code__icontains=filters["doctor"])
        )

    # dərmanların toplanması
    for prescription in prescriptions:
        doctor = prescription.doctor
        if not doctor:
            continue

        summary = doctor_summary[doctor.id]

        for item in prescription.items.all():
            name = item.drug.ad
            summary["drugs"][name] = summary["drugs"].get(name, 0) + item.quantity
            summary["total_quantity"] += item.quantity

    # Payments
    doctor_ids = list(doctor_summary.keys())

    if doctor_ids:
        payments = DoctorPayment.objects.filter(doctor_id__in=doctor_ids)

        if filters["region"]:
            payments = payments.filter(region_id=filters["region"])
        if month_int:
            payments = payments.filter(date__month=month_int)
        if year_int:
            payments = payments.filter(date__year=year_int)

        for payment in payments:
            doctor_summary[payment.doctor_id]["payments"][payment.payment_type] += payment.amount

    # yekunu hesablamaq
    for summary in doctor_summary.values():
        evvelki = summary["evvelki_borc"]
        avans = summary["payments"]["avans"]
        investisiya = summary["payments"]["investisiya"]
        geri = summary["payments"]["geriqaytarma"]
        datasiya = summary["datasiya"]
        silinen = summary["silinen_miqdar"]

        summary["yekun_borc"] = evvelki + avans + investisiya + datasiya + geri - silinen

    doctor_rows = list(doctor_summary.values())
    drugs = Drug.objects.filter(is_active=True).order_by("ad")

    return filters, doctor_rows, drugs
# -----------------------------
# MAIN VIEW
# -----------------------------

def monthly_reports(request):
    """Monthly doctor performance + financial report."""
    today = date.today()
    filters, doctor_rows, drugs = prepare_reports_data(request)

    regions = Region.objects.order_by("name")
    years = range(today.year - 2, today.year + 3)

    stats = {
        "doctor_count": len(doctor_rows),
        "total_debt": sum(row["yekun_borc"] for row in doctor_rows),
        "total_prescriptions": sum(row["total_quantity"] for row in doctor_rows),
    }

    table_info = build_table_info(filters, regions, stats)

    return render(
        request,
        "prescriptions/reports.html",
        {
            "filters": filters,
            "regions": regions,
            "drugs": drugs,
            "doctor_rows": doctor_rows,
            "years": years,
            "months": MONTH_CHOICES,
            "stats": stats,
            "table_info": table_info,
        }
    )


# -----------------------------
# CLOSE MONTH (HESABATI BAĞLA)
# -----------------------------

@transaction.atomic
def close_month_report(request):
    """
    Close current filtered month:
    - If month/year already closed -> error.
    - Otherwise:
        1) Read live report data via prepare_reports_data
        2) Save snapshot to MonthlyDoctorReport
        3) Move doctors into next month:
           evvelki_borc = yekun_borc, others reset to 0.
    """
    today = date.today()
    month = safe_int(request.GET.get("month"), today.month)
    year = safe_int(request.GET.get("year"), today.year)
    region = (request.GET.get("region") or "").strip()

    if not month or not year:
        messages.error(request, "Month and year are required to close report.")
        return redirect("reports:list")
    
    if not region:
        messages.error(request, "Bölgə seçilməlidir.")
        return redirect("reports:list")

    # Month cannot be closed twice
    if MonthlyDoctorReport.objects.filter(month=month, year=year).exists():
        messages.error(request, "Bu ay üçün hesabat artıq bağlanıb.")
        return redirect("reports:list")

    # Always use live data for closing (ignore any existing snapshot logic)
    filters, doctor_rows, drugs = prepare_reports_data(request)

    # Create snapshot per doctor
    for row in doctor_rows:
        MonthlyDoctorReport.objects.create(
        doctor=row["doctor"],
        region=row["doctor"].region,
        year=year,
        month=month,

        drugs_data=row["drugs"],    # ✔ models.JSONField → drugs_data
        total_quantity=row["total_quantity"],

        evvelki_borc=row["evvelki_borc"],
        hesablanan=row["hesablanan"],
        silinen_miqdar=row["silinen_miqdar"],   # ✔ model field

        avans=row["payments"]["avans"],
        investisiya=row["payments"]["investisiya"],
        geriqaytarma=row["payments"]["geriqaytarma"],
        datasiya=row["datasiya"],

        yekun_borc=row["yekun_borc"],
    )


    # Reset doctors for next month
    for row in doctor_rows:
        doctor = row["doctor"]
        doctor.evvelki_borc = row["yekun_borc"]
        doctor.hesablanmish_miqdar = Decimal("0")
        doctor.silinen_miqdar = Decimal("0")
        doctor.datasiya = Decimal("0")
        doctor.save(update_fields=[
            "evvelki_borc",
            "hesablanmish_miqdar",
            "silinen_miqdar",
            "datasiya",
        ])

    messages.success(request, "Hesabat uğurla bağlandı.")
    return redirect("reports:list")


def export_reports_excel(request):
    """Export monthly reports to Excel file."""
    filters, doctor_rows, drugs = prepare_reports_data(request)

    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Aylıq Həkim Hesabatı"

    # Define styles
    header_font = Font(name='Calibri', size=11, bold=True, color='ffffff')
    header_fill = PatternFill(start_color='7EA6E0', end_color='7EA6E0', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    cell_font = Font(name='Calibri', size=10)
    cell_alignment = Alignment(horizontal='left', vertical='center')
    center_alignment = Alignment(horizontal='center', vertical='center')
    number_alignment = Alignment(horizontal='right', vertical='center')

    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # Build headers
    headers = [
        '№',
        'Bölgə',
        'Kod',
        'Həkim',
        'İxtisas',
        'Kateqoriya',
        'Dərəcə',
        'Əvvəlki Borc',
    ]

    # Add drug columns
    for drug in drugs:
        headers.append(drug.ad)

    # Add remaining columns
    headers.extend([
        'Total',
        'Hesablanan',
        'Silinən',
        'Avans',
        'İnvestisiya',
        'Geri Qaytarma',
        'Datasiya',
        'Yekun',
    ])

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Set column widths
    ws.column_dimensions['A'].width = 5   # №
    ws.column_dimensions['B'].width = 15  # Bölgə
    ws.column_dimensions['C'].width = 12  # Kod
    ws.column_dimensions['D'].width = 25  # Həkim
    ws.column_dimensions['E'].width = 15  # İxtisas
    ws.column_dimensions['F'].width = 12  # Kateqoriya
    ws.column_dimensions['G'].width = 10  # Dərəcə
    ws.column_dimensions['H'].width = 15  # Əvvəlki Borc

    # Drug columns
    drug_start_col = 9
    for idx, drug in enumerate(drugs):
        col_letter = ws.cell(row=1, column=drug_start_col + idx).column_letter
        ws.column_dimensions[col_letter].width = 12

    # Remaining columns
    remaining_cols = ['Total', 'Hesablanan', 'Silinen', 'Avans', 'Investisiya', 'Geri Qaytarma', 'Datasiya', 'Yekun']
    remaining_start_col = drug_start_col + len(drugs)
    for idx, col_name in enumerate(remaining_cols):
        col_letter = ws.cell(row=1, column=remaining_start_col + idx).column_letter
        ws.column_dimensions[col_letter].width = 15

    # Write data rows
    for row_num, row in enumerate(doctor_rows, 2):
        col_num = 1

        # №
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = row_num - 1
        cell.font = cell_font
        cell.alignment = center_alignment
        cell.border = border
        col_num += 1

        # Bölgə
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = row['doctor'].region.name if row['doctor'].region else '-'
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = border
        col_num += 1

        # Kod
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = row['doctor'].code
        cell.font = cell_font
        cell.alignment = center_alignment
        cell.border = border
        col_num += 1

        # Həkim
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = row['doctor'].ad
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = border
        col_num += 1

        # İxtisas
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = row['doctor'].ixtisas.name if row['doctor'].ixtisas else '-'
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = border
        col_num += 1

        # Kateqoriya
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = row['doctor'].category
        cell.font = cell_font
        cell.alignment = center_alignment
        cell.border = border
        col_num += 1

        # Dərəcə
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = row['doctor'].get_degree_display()
        cell.font = cell_font
        cell.alignment = center_alignment
        cell.border = border
        col_num += 1

        # Əvvəlki Borc
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = float(row['evvelki_borc'])
        cell.number_format = '#,##0.00'
        cell.font = cell_font
        cell.alignment = number_alignment
        cell.border = border
        col_num += 1

        # Drug columns
        for drug in drugs:
            cell = ws.cell(row=row_num, column=col_num)
            drug_count = row['drugs'].get(drug.ad, 0)
            cell.value = drug_count
            cell.font = cell_font
            cell.alignment = center_alignment
            cell.border = border
            col_num += 1

        # Total
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = row['total_quantity']
        cell.font = cell_font
        cell.alignment = center_alignment
        cell.border = border
        col_num += 1

        # Hesablanan
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = float(row['hesablanan'])
        cell.number_format = '#,##0.00'
        cell.font = cell_font
        cell.alignment = number_alignment
        cell.border = border
        col_num += 1

        # Silinen
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = float(row['silinen_miqdar'])
        cell.number_format = '#,##0.00'
        cell.font = cell_font
        cell.alignment = number_alignment
        cell.border = border
        col_num += 1

        # Avans
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = float(row['payments']['avans'])
        cell.number_format = '#,##0.00'
        cell.font = cell_font
        cell.alignment = number_alignment
        cell.border = border
        col_num += 1

        # Investisiya
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = float(row['payments']['investisiya'])
        cell.number_format = '#,##0.00'
        cell.font = cell_font
        cell.alignment = number_alignment
        cell.border = border
        col_num += 1

        # Geri Qaytarma
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = float(row['payments']['geriqaytarma'])
        cell.number_format = '#,##0.00'
        cell.font = cell_font
        cell.alignment = number_alignment
        cell.border = border
        col_num += 1

        # Datasiya
        cell = ws.cell(row=row_num, column=col_num)
        datasiya_value = row['datasiya']
        if isinstance(datasiya_value, date):
            cell.value = float(0)  # If it's a date, treat as 0
        else:
            cell.value = float(datasiya_value)
        cell.number_format = '#,##0.00'
        cell.font = cell_font
        cell.alignment = number_alignment
        cell.border = border
        col_num += 1

        # Yekun
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = float(row['yekun_borc'])
        cell.number_format = '#,##0.00'
        cell.font = cell_font
        cell.alignment = number_alignment
        cell.border = border

    # Generate filename
    month_label = next((label for val, label in MONTH_CHOICES if val == filters["month"]), "Bütün aylar")
    filename = f"Ayliq_Hekim_Hesabati_{month_label}_{filters['year']}.xlsx"

    # Create HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response






