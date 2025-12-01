from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from .models import Doctor, DoctorPayment
from regions.models import Region, City, Clinic, Specialization
from subscription.decorators import subscription_required
from prescriptions.models import Prescription
from django.http import JsonResponse


@login_required
@subscription_required
def doctor_list(request):
    """
    List all doctors from this company's database with pagination.
    Shows 25 doctors per page for optimal performance.
    """
    from decimal import Decimal, InvalidOperation
    
    # Get all doctors with related data
    # We'll load them one by one to catch Decimal errors
    try:
        doctor_ids = Doctor.objects.values_list('id', flat=True).order_by('-created_at')
        
        # Safely load doctors and handle Decimal fields
        doctors_safe = []
        for doctor_id in doctor_ids:
            try:
                doctor = Doctor.objects.select_related(
                    'region', 'city', 'clinic', 'ixtisas'
                ).get(id=doctor_id)
                
                # Safely convert Decimal fields
                def safe_decimal(value, default=Decimal('0.00')):
                    if value is None:
                        return default
                    try:
                        if isinstance(value, str):
                            return Decimal(str(value).replace(',', '.'))
                        return Decimal(str(value))
                    except (InvalidOperation, ValueError, TypeError):
                        return default
                
                # Create a wrapper object that safely handles Decimal fields
                class SafeDoctor:
                    def __init__(self, doctor):
                        self._doctor = doctor
                        self.id = doctor.id
                        self.code = doctor.code
                        self.ad = doctor.ad
                        self.telefon = doctor.telefon
                        self.email = doctor.email
                        self.gender = doctor.gender
                        self.category = doctor.category
                        self.degree = doctor.degree
                        self.is_active = doctor.is_active
                        self.created_at = doctor.created_at
                        self.datasiya = doctor.datasiya
                        self.region = doctor.region
                        self.city = doctor.city
                        self.clinic = doctor.clinic
                        self.ixtisas = doctor.ixtisas
                        self.evvelki_borc = safe_decimal(doctor.evvelki_borc)
                        self.hesablanmish_miqdar = safe_decimal(doctor.hesablanmish_miqdar)
                        self.silinen_miqdar = safe_decimal(doctor.silinen_miqdar)
                        self.yekun_borc = safe_decimal(doctor.yekun_borc)
                    
                    def get_gender_display(self):
                        return self._doctor.get_gender_display()
                
                doctors_safe.append(SafeDoctor(doctor))
            except Exception:
                # Skip problematic doctors
                continue
        
        # Pagination - 25 doctors per page
        paginator = Paginator(doctors_safe, 25)
        page = request.GET.get('page', 1)
        
        try:
            doctors = paginator.page(page)
        except PageNotAnInteger:
            doctors = paginator.page(1)
        except EmptyPage:
            doctors = paginator.page(paginator.num_pages)
            
    except Exception as e:
        # If there's an error, return empty list
        paginator = Paginator([], 25)
        doctors = paginator.page(1)
        doctors_safe = []
    
    # Get filter options
    regions = Region.objects.all().order_by('name')
    specializations = Specialization.objects.all().order_by('name')
    clinics = Clinic.objects.filter(is_active=True).order_by('name')
    
    # Get total count for display
    total_doctors = len(doctors_safe)
    
    context = {
        'doctors': doctors,
        'total_doctors': total_doctors,
        'regions': regions,
        'specializations': specializations,
        'clinics': clinics,
    }
    
    return render(request, 'doctors/list.html', context)


@login_required
@subscription_required
def add_doctor(request):
    """Add new doctor with all required fields"""
    company = request.company
    
    # Check if company reached doctor limit
    current_doctors = Doctor.objects.count()
    if current_doctors >= company.max_doctors:
        messages.error(
            request,
            f'Planınızın {company.max_doctors} həkim limitinə çatdınız. '
            'Daha çox əlavə etmək üçün planınızı yüksəldin.'
        )
        return redirect('subscription:plans')
    
    if request.method == 'POST':
        # Get form data
        ad = request.POST.get('ad', '').strip()
        region_id = request.POST.get('region')
        city_id = request.POST.get('city')
        ixtisas_id = request.POST.get('ixtisas')
        degree = request.POST.get('degree')
        category = request.POST.get('category')
        telefon = request.POST.get('telefon', '').strip()
        gender = request.POST.get('gender')
        clinic_id = request.POST.get('clinic')
        
        # Validation - Required fields
        errors = []
        if not ad:
            errors.append('Ad Soyad tələb olunur!')
        if not region_id:
            errors.append('Bölgə seçilməlidir!')
        if not ixtisas_id:
            errors.append('İxtisas seçilməlidir!')
        if not degree:
            errors.append('Dərəcə seçilməlidir!')
        if not category:
            errors.append('Kateqoriya seçilməlidir!')
        if not gender:
            errors.append('Cinsiyyət seçilməlidir!')
        
        if errors:
            for error in errors:
                messages.error(request, error)
            context = get_add_doctor_context(current_doctors, company.max_doctors)
            return render(request, 'doctors/add.html', context)
        
        # Create doctor
        try:
            doctor_data = {
                'ad': ad,
                'region_id': region_id,
                'ixtisas_id': ixtisas_id,
                'degree': degree,
                'category': category,
                'gender': gender,
            }
            
            # Optional fields
            if city_id:
                doctor_data['city_id'] = city_id
            if telefon:
                doctor_data['telefon'] = telefon
            if clinic_id:
                doctor_data['clinic_id'] = clinic_id
            
            Doctor.objects.create(**doctor_data)
            messages.success(request, 'Həkim uğurla əlavə edildi!')
            return redirect('doctors:list')
            
        except Exception as e:
            messages.error(request, f'Xəta baş verdi: {str(e)}')
            context = get_add_doctor_context(current_doctors, company.max_doctors)
            return render(request, 'doctors/add.html', context)
    
    # GET request - show form
    context = get_add_doctor_context(current_doctors, company.max_doctors)
    return render(request, 'doctors/add.html', context)


def get_add_doctor_context(current_doctors, max_doctors):
    """Helper function to get context for add doctor form"""
    return {
        'doctors_count': current_doctors,
        'doctors_limit': max_doctors,
        'remaining': max_doctors - current_doctors,
        'regions': Region.objects.all().order_by('name'),
        'cities': City.objects.select_related('region').all().order_by('name'),
        'clinics': Clinic.objects.select_related('city', 'city__region').filter(is_active=True).order_by('name'),
        'specializations': Specialization.objects.all().order_by('name'),
        'genders': Doctor.GENDER_CHOICES,
        'categories': Doctor.CATEGORY_CHOICES,
        'degrees': Doctor.DEGREE_CHOICES,
    }


@login_required
@subscription_required
def doctor_detail(request, doctor_id):
    """
    Display detailed information about a specific doctor
    """
    prescriptions = Prescription.objects.filter(doctor=doctor_id)
    try:
        doctor = Doctor.objects.select_related(
            'region',
            'city',
            'clinic',
            'ixtisas'
        ).get(id=doctor_id)
    except Doctor.DoesNotExist:
        messages.error(request, 'Həkim tapılmadı.')
        return redirect('doctors:list')
    
    context = {
        'doctor': doctor,
        'prescriptions': prescriptions,
    }
    
    return render(request, 'doctors/detail.html', context)


@login_required
@subscription_required
def export_doctors_excel(request):
    """
    Export all doctors to Excel file with formatting
    """
    # Get all doctors for the current company
    doctors = Doctor.objects.select_related(
        'region',
        'city',
        'clinic',
        'ixtisas'
    ).all().order_by('code')
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Həkimlər"
    
    # Define styles
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    cell_font = Font(name='Arial', size=11)
    cell_alignment = Alignment(horizontal='left', vertical='center')
    number_alignment = Alignment(horizontal='right', vertical='center')
    
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Define column headers
    headers = [
        ('Kod', 12),
        ('Ad Soyad', 25),
        ('İxtisas', 20),
        ('Bölgə', 15),
        ('Şəhər', 15),
        ('Klinika', 30),
        ('Telefon', 18),
        ('Email', 25),
        ('Cinsiyyət', 12),
        ('Kateqoriya', 12),
        ('Dərəcə', 10),
        ('Əvvəlki Borc', 15),
        ('Hesablanmış', 15),
        ('Silinən', 15),
        ('Yekun Borc', 15),
        ('Qeydiyyat Tarixi', 18),
        ('Status', 12),
    ]
    
    # Write headers
    for col_num, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = width
    
    # Write data
    for row_num, doctor in enumerate(doctors, 2):
        # Kod
        cell = ws.cell(row=row_num, column=1)
        cell.value = doctor.code
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = border
        
        # Ad Soyad
        cell = ws.cell(row=row_num, column=2)
        cell.value = doctor.ad
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = border
        
        # İxtisas
        cell = ws.cell(row=row_num, column=3)
        cell.value = doctor.ixtisas.name if doctor.ixtisas else '-'
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = border
        
        # Bölgə
        cell = ws.cell(row=row_num, column=4)
        cell.value = doctor.region.name if doctor.region else '-'
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = border
        
        # Şəhər
        cell = ws.cell(row=row_num, column=5)
        cell.value = doctor.city.name if doctor.city else '-'
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = border
        
        # Klinika
        cell = ws.cell(row=row_num, column=6)
        cell.value = doctor.clinic.name if doctor.clinic else '-'
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = border
        
        # Telefon
        cell = ws.cell(row=row_num, column=7)
        cell.value = doctor.telefon
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = border
        
        # Email
        cell = ws.cell(row=row_num, column=8)
        cell.value = doctor.email or '-'
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = border
        
        # Cinsiyyət
        cell = ws.cell(row=row_num, column=9)
        cell.value = doctor.get_gender_display()
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = border
        
        # Kateqoriya
        cell = ws.cell(row=row_num, column=10)
        cell.value = doctor.category
        cell.font = cell_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        
        # Dərəcə
        cell = ws.cell(row=row_num, column=11)
        cell.value = doctor.degree
        cell.font = cell_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        
        # Əvvəlki Borc
        cell = ws.cell(row=row_num, column=12)
        cell.value = float(doctor.evvelki_borc)
        cell.number_format = '#,##0.00'
        cell.font = cell_font
        cell.alignment = number_alignment
        cell.border = border
        
        # Hesablanmış Miqdar
        cell = ws.cell(row=row_num, column=13)
        cell.value = float(doctor.hesablanmish_miqdar)
        cell.number_format = '#,##0.00'
        cell.font = cell_font
        cell.alignment = number_alignment
        cell.border = border
        
        # Silinən Miqdar
        cell = ws.cell(row=row_num, column=14)
        cell.value = float(doctor.silinen_miqdar)
        cell.number_format = '#,##0.00'
        cell.font = cell_font
        cell.alignment = number_alignment
        cell.border = border
        
        # Yekun Borc
        cell = ws.cell(row=row_num, column=15)
        cell.value = float(doctor.yekun_borc)
        cell.number_format = '#,##0.00'
        cell.font = Font(name='Arial', size=11, bold=True)
        cell.alignment = number_alignment
        cell.border = border
        
        # Color code based on debt
        if doctor.yekun_borc > 0:
            cell.font = Font(name='Arial', size=11, bold=True, color='FF0000')  # Red for debt
        elif doctor.yekun_borc < 0:
            cell.font = Font(name='Arial', size=11, bold=True, color='00FF00')  # Green for credit
        
        # Qeydiyyat Tarixi
        cell = ws.cell(row=row_num, column=16)
        cell.value = doctor.datasiya.strftime('%d.%m.%Y') if doctor.datasiya else '-'
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = border
        
        # Status
        cell = ws.cell(row=row_num, column=17)
        cell.value = 'Aktiv' if doctor.is_active else 'Deaktiv'
        cell.font = cell_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        
        # Color code status
        if doctor.is_active:
            cell.fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
        else:
            cell.fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
    
    # Freeze first row
    ws.freeze_panes = 'A2'
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Generate filename with company name and date
    company_name = request.company.name if hasattr(request, 'company') else 'Company'
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'Həkimlər_{company_name}_{timestamp}.xlsx'
    
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save workbook to response
    wb.save(response)
    
    return response

@login_required
@subscription_required
def get_doctors_by_region(request):
    region_id = request.GET.get("region_id")
    if not region_id:
        return JsonResponse({'error': 'region_id tələb olunur'}, status=400)

    doctors = Doctor.objects.filter(region_id=region_id, is_active=True).values("id", "ad").order_by("ad")
    payload = [{'id': doctor['id'], 'name': doctor['ad']} for doctor in doctors]
    return JsonResponse(payload, safe=False)

@login_required
@subscription_required
def add_doctor_payment(request):
    regions = Region.objects.all()

    if request.method == "POST":
        region_id = request.POST.get("region")
        doctor_id = request.POST.get("doctor")
        payment_type = request.POST.get("payment_type")
        amount = request.POST.get("amount")
        date = request.POST.get("date")

        if not region_id or not doctor_id or not payment_type or not amount or not date:
            messages.error(request, "Zəhmət olmasa bütün xanaları doldurun.")
        else:
            DoctorPayment.objects.create(
                region_id=region_id,
                doctor_id=doctor_id,
                payment_type=payment_type,
                amount=amount,
                date=date
            )
            messages.success(request, "Ödəniş uğurla əlavə edildi!")
            return redirect('doctors:add_payment')

           

    return render(request, "doctors/add-payment.html", {
        "regions": regions
    })
